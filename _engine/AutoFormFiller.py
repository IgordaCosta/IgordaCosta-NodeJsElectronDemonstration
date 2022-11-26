# -*- coding: utf-8 -*-
"""
Created on Thu Jul 11 19:47:41 2019

@author: IgorDC
"""

import pandas

import numpy as np

import sys

import os

# from StyleFrame import StyleFrame, Styler
from styleframe import StyleFrame, Styler 

from win32com.client import Dispatch

import win32com.client as win32

import openpyxl

# from PyQt5 import QtCore, QtGui, QtWidgets

# from PyQt5.QtGui import QClipboard

# from PyQt5.QtWidgets import QMainWindow, QApplication, QAction, QFileDialog, QTableWidgetItem

import glob

import shutil

import sqlite3

from sqlalchemy import create_engine

import re

import math

#from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

#from PyQt5 import QtGui, QtWidgets
# =============================================================================
# from PyQt5.QtCore import QObject
# from PyQt5.QtGui import QPainter, QFont, QColor, QPen
# =============================================================================

# from AutoFormFillerMainEN import *

# from AutoFormFillerJobNameEN import *

#from Trabalho_Para_Nilson_Henrique_Inicio_Salve_Documento import *

#from Trabalho_Para_Nilson_Henrique_Adi_Information import *

# from AutoFormFiller_Adi_InformationEN import *

# from AutoFormFillerOpConcluidaEN import *

# from AutoFormFillerCopyandPasteMarkEN import *

# from AutoFormFiller_Inicio_Salve_DocumentoEN import *

# from AutoFormFiller_Add_Info_SaveEN import *

# from AutoFormFillerFileExistsEN import *

# from AutoFormFiller_Add_Info_TabelaEN import *

# from AutoFormFillerJobNameNotExistEN import *

# from AutoFormFillerAreYouSureDeleteEN import *

# from AutoFormFillerCloseDoc2ContinueEN import *

# from AutoFormFillerProhibitedFileNamesUsedEN import *







class autoFormFiller(QMainWindow):
    
    def __init__(self):
        self.Main()
        
    def Main(self):
        super().__init__()
        
        
        self.SqlDatase=True
        self.UseExcel=False
        
        
        self.FilesInDatabaseLocation='FilesInDatabase.xlsx'
        
        if self.SqlDatase==True:
             self.FilesInDatabaseLocation= self.FilesInDatabaseLocation.split(".")[0]
        
        self.columnsFileinDatabase=['Job Name', 'File KEY', 'File Saved Location']
        
        self.DatabaseName="AutoFormFiller.db"
        
        SqlDatase=self.SqlDatase
        
        # if not QtWidgets.QApplication.instance():
        #      app = QtWidgets.QApplication(sys.argv)
        # else:
        #      app = QtWidgets.QApplication.instance() 
        
        # MainWindow = QtWidgets.QMainWindow()
        
        

        # self.ui0 = Ui_MainWindow()
        # self.ui0.setupUi(MainWindow)
        
        self.ChangeStartupDirectory(Folder='AutoFormFillerKey')
        
        self.uploadDatabase(SqlDatase=SqlDatase)
        
        #self.ui.AddFile.clicked.connect(self.clickedGetFileName)
        
        # self.ui0.AddFile.clicked.connect(self.clickedGetFileName)
        
        # self.ui0.DeleteFile.clicked.connect(self.DeleteItemDatabase)
        
        # self.ui0.LoadToMulFiles.clicked.connect(self.FuncLoadToMultipleFiles)
        
        # self.ui0.LoadToSingleFile.clicked.connect(self.FuncLoadToSingleFile)
        
        
        # self.InitialValue=False
        # self.ui0.tableWidget.cellClicked.connect(self.cellClicked)
        # self.ui0.tableWidget.currentCellChanged.connect(self.cellClicked)
        
        
        
        
        
        
        
        
        
        
        
        MainWindow.show()
        sys.exit(app.exec_())
        
        
        
        
        

# =============================================================================
#         super().__init__()
#         self.ui = Ui_MainWindow()
#         self.ui.setupUi(self)
#         
#         #app = QtWidgets.QApplication(sys.argv)
#         
#         
#         if not QtWidgets.QApplication.instance():
#             app = QtWidgets.QApplication(sys.argv)
#         else:
#             app = QtWidgets.QApplication.instance() 
#         
#         
#         
#         
#         
#         
#         MainWindow = QtWidgets.QMainWindow()
#         
#         
#         
#         
#         self.ui = Ui_MainWindow()
#         self.ui.setupUi(MainWindow)
#         
# # =============================================================================
# #         self.ChangeStartupDirectory(Folder='AutoFormFillerKey')
# # =============================================================================
#         
#         
#         
#         #self.ui.AddFile.clicked.connect(self.clickedGetFileName)
#         #self.ui.BCadUmCliente.clicked.connect(self.clickedCadClient)
#         
#         
#         
#         MainWindow.show()
#         sys.exit(app.exec_())
# =============================================================================
    
        
    def cellClicked(self,row,col):
        #print(row,col)
        self.InitialValue=True
        self.rowClicked=row
        
        print(self.rowClicked)
        
    def OptionForItemInDatabase(self, option):
        
        SqlDatase=self.SqlDatase
        FilesInDatabaseLocation=self.FilesInDatabaseLocation
        columnsFileinDatabase=self.columnsFileinDatabase
        
        if SqlDatase==True:
            FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
        
        if self.InitialValue==False:
            Redowhile=True
            rsp=1
            datafillName=''
            while (Redowhile==True or datafillName =='')  and rsp==1:
                rsp, datafillName=self.JobNameWindow()
                
                datafillName=datafillName.split()
                datafillName=" ".join(datafillName)
                
                if SqlDatase==False:
                    ####Added part below
                    FilesInDatabaseExists=False
                    try:
                        FilesInDatabase=pandas.read_excel(FilesInDatabaseLocation)
                        FilesInDatabaseExists=True
                    except FileNotFoundError:
                        print('database DOES NOT exist... continue')
                        FilesInDatabaseExists=False
                        
                if SqlDatase==True:
                    FilesInDatabase , FilesInDatabaseExists = self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
                    
                if FilesInDatabaseExists==True:
                    
                    if datafillName not in list(FilesInDatabase[columnsFileinDatabase[0]]) and datafillName !='':
                        Redowhile=True
                        print("message to choose a valida option")
                        self.DJobNameNotExit()
                        
                    else:
                       Redowhile=False 
                
                if FilesInDatabaseExists==False:
                    print("there is no data in database")
            
            ####Added part above
            
            if rsp==1 and datafillName !='':

                if option=="delete":
                    self.DeleteFromDataFillName(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase, SqlDatase=SqlDatase)
                if option =="LoadToMulFiles":

                    self.LoadToMultipleFilesDetails(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
                    pass
                if option =="LoadToSingleFile":
                    
                        self.LoadToSingleFileDetails(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
                
            pass
        
        print("self.rowClicked before none check")
        if self.InitialValue==True:
            self.InitialValue=False
            
            try:
                datafillName = self.ui0.tableWidget.item(self.rowClicked, 0).text()
                Skip=False
            except AttributeError:
                print("Nontype object tribute error")
                Skip=True
                
            if Skip==False:
                
                if option=="delete":
                    rsp=self.DAreYouSureDelete(datafillName=datafillName)
                if option =="LoadToMulFiles" or option =="LoadToSingleFile":
                     rsp=self.DloadItens(datafillName=datafillName)
                     
                if rsp==1:
                    if option=="delete":
                        self.DeleteFromDataFillName(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase, SqlDatase=SqlDatase)
                    if option =="LoadToMulFiles":
                        self.LoadToMultipleFilesDetails(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
                        pass
                    if option =="LoadToSingleFile":
                        self.LoadToSingleFileDetails(datafillName=datafillName, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
        
        
        
    def DeleteItemDatabase(self):
        self.OptionForItemInDatabase(option="delete")
        
        
    def DeleteFromDataFillName(self, datafillName, FilesInDatabaseLocation, columnsFileinDatabase, SqlDatase):
        self.dropSqlTable(database=self.DatabaseName, table="KEY_"+datafillName)
    
        databaseUsed0=self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
        databaseUsed=databaseUsed0[0]
        
        databaseUsed=databaseUsed[~(databaseUsed[columnsFileinDatabase[0]]==datafillName)]
        
        pandas.DataFrame.reset_index(databaseUsed, inplace=True)
        databaseUsed=databaseUsed[columnsFileinDatabase]
        print(databaseUsed)
        self.writeToSqlite(frame=databaseUsed,table_name=FilesInDatabaseLocation)
        self.uploadDatabase(SqlDatase=SqlDatase)
        self.OperaConcluidaWindow()
        
    def FuncLoadToSingleFile(self):
        self.OptionForItemInDatabase(option="LoadToSingleFile")
        
        pass
    
    def FolderHasProribitedFileNames(self):
        
        filename="OneFileToMultipleFiles.xlsx"
        
        location="AutoFormFillerKey"
        
        home = os.path.expanduser('~')    
        
        filePath = os.path.join(home, 'Documents',location)
        
        filename2=filePath+"\\"+filename
        
        #filename="OneFileToMultipleFiles.xlsx"
        
        #home = os.path.expanduser('~')
                
        print(home)
        
        location = os.path.join(home, 'Desktop')
        
        SelectedFolder=''
        
        ExistInside=True
        rsp=1
        self.IgonoreClicked=False
        while ExistInside==True and rsp==1:
            SelectedFolder = QFileDialog.getExistingDirectory(self, str("Select the folder where the files are located"),location, QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
            
            print(SelectedFolder)
                   
             
            SelectedFolderFiles=self.GetExcelFilesinLocation(location=SelectedFolder, IncludeFolder=True)
            
            print(SelectedFolderFiles)
            
            
            filesInFolder1=self.GetExcelFilesinLocation(location="AutoFormFillerFiles", IncludeFolder=True)
            
            listTocheck=[filename2]+filesInFolder1
            
            print(listTocheck)
            print("listTocheck before check")
            
            
            print(SelectedFolderFiles)
            print("SelectedFolderFiles before check")
            
            
            #ExistInside, ListofInlist=self.valueInList(firstList=SelectedFolderFiles, listTocheck=listTocheck)
            ExistInside, ListofInlist=self.valueInList(firstList=SelectedFolderFiles,listTocheck=listTocheck, InListFlag=True, KeepOriginal=False)
            
            print(ExistInside)
            print(ListofInlist)
            
            print("ListofInlist after check")
            
            ListofInlist2=self.ListToSentence(ListUsed=ListofInlist)
            
            print(ListofInlist2)
            if ExistInside==True:
                print("ok its inside")
                rsp=self.DProriFileNamesInFolderUsed(filenameUsed=ListofInlist2)
                print(rsp)
        
        if self.IgonoreClicked==True:
            rsp=1
            ExistInside, SelectedFolderFiles=self.valueInList(firstList=listTocheck,listTocheck=SelectedFolderFiles, InListFlag=False, KeepOriginal=True)
            print(SelectedFolderFiles)
            print("SelectedFolderFiles after igore button clicked")
# =============================================================================
#             print(len(SelectedFolderFiles))
#             for value in ListofInlist:
#                 SelectedFolderFiles.remove(value)
#             print(SelectedFolderFiles)
#             print(len(SelectedFolderFiles))
# =============================================================================
            
       
        return SelectedFolderFiles, rsp
    
    
    
    def LoadToSingleFileDetails(self, datafillName, FilesInDatabaseLocation, columnsFileinDatabase):
        
        ####TODO add a filename column to the final database and insert all the filenames for each row in it
        
# =============================================================================
#         print("start Load to Single File Details")
#         
#         home = os.path.expanduser('~')
#         
#         print(home)
#         
#         location = os.path.join(home, 'Desktop')
#         
#         SelectedFolder=''
#         
#         SelectedFolder = QFileDialog.getExistingDirectory(self, str("Select the folder where the files are located"),location, QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
#         
#         print(SelectedFolder)
#         
#         if SelectedFolder !='':
# =============================================================================
        
        
        SelectedFolderFiles, rsp= self.FolderHasProribitedFileNames()
        
        print(SelectedFolderFiles)
        
        if len(SelectedFolderFiles)!=0 and rsp==1:
            print("folder chosen and it has data")
            
            FileFolder='AutoFormFillerSingleFileOutput'
            filename=''
            
            InFileLocation=True
            rsp=1
            self.IgonoreClicked=False
            while InFileLocation==True and rsp==1:
                rsp, filename=self.NameFileToSaveAs()
                if rsp==1:
                    if filename !='':
                        updateNamesFile=filename.split('.')[0]
                        FinalFilename=updateNamesFile+".xlsx"
                        InFileLocation, SenteceOpenFiles=self.CheckIfInFileLocation(fileToCheck=FinalFilename, location=FileFolder)
                    
                    else:
                        InFileLocation=True
                    
                    if InFileLocation==True:
                        rsp=self.DSingleFileNameUsed(filenameUsed=FinalFilename)
            if self.IgonoreClicked:
                rsp=1
                InFileLocation=False
            
            
            if rsp==1:
   
                print(InFileLocation, "<<<InFileLocation value")
                if InFileLocation==False:
                    home = os.path.expanduser('~')
                    
                    filePath = os.path.join(home, 'Documents',FileFolder)
                     
                    if not os.path.exists(filePath):
                        os.makedirs(filePath)
                    
                    updateNamesFile=filename.split('.')[0]
                    FinalFilename=updateNamesFile+".xlsx"
                    
                    print("FinalFilename location name")
                    print("filename used in the end")
                    
                    #filesInFolder=glob.glob(SelectedFolder+"\\*.xl*")
                    filesInFolder=SelectedFolderFiles
                    print(filesInFolder)
                    
                    numberOfFilesinFolder=len(filesInFolder)
                    print(numberOfFilesinFolder)
                    
                    databaseUsed0=self.readSqlDatabase(table_name="KEY_"+datafillName)
                    databaseUsed=databaseUsed0[0]
                    
                    print(databaseUsed)
                    
                    print("databaseUsed")
                    
                    itemAddress=databaseUsed.columns.values[1:]
                    print(itemAddress)
                    
                    itemDescription=databaseUsed.values[0][1:]
                    print(itemDescription)
                    
                    rowList, columnList=self.getAddressesFromColumn(itemAddress=itemAddress)
                    
                    numberOfRows=numberOfFilesinFolder
                    print(numberOfRows)
                    
                    ListUsed=[]
                    for rw in range(numberOfRows):
                        print("open file in row to extract data")
                        ListUsed.append([])
                        
                        xlfile=filesInFolder[rw]
                        
                        wb =self.openWorkbook(xlfile)
                        ws = wb.Sheets(1)
                        
                        if wb!=None:
                        
                            for cl in range(1,len(itemAddress)+1):
                # =============================================================================
                #                 print(storedValues.values[rw][cl])
                #                 #print(type(storedValues.values[rw][cl]))
                #                 print(str(storedValues.values[rw][cl]).upper()=="NAN")
                # =============================================================================
                                i=rowList[cl-1]
                                c=columnList[cl-1]
                                
                                ValueSaved=ws.Cells(i,c).Value
                                
                                ListUsed[rw].append(ValueSaved)
                            
                            wb.Close(True)
                   
                    print(ListUsed)
                    print("ListUsed values at the end")
                    
                    
                    df = pandas.DataFrame(ListUsed, columns =itemDescription)
                    df["Save Name"]=filesInFolder
                    
                    
                    itemDescription=["Save Name"]+list(itemDescription)
                    
                    df=df[itemDescription]
                    
                    
                    
                    print(df)
                    print("dataframe final")
                    
                    
# =============================================================================
#                     filePath = os.path.join(home, 'Documents','AutoFormFillerSingleFileOutput')
#                         
#                         
#                     if not os.path.exists(filePath):
#                         os.makedirs(filePath)
#                     
#                     
#                     self.ChangeStartupDirectory(Folder='AutoFormFillerSingleFileOutput')
# =============================================================================
                    
                    
                    #filename="testing.xlsx"
                    
                    
                    
                    self.ChangeStartupDirectory(Folder=FileFolder)
                    FileSaved=False 
                    rsp=1
                    
                    while FileSaved==False and rsp==1:
                        print(rsp)
                        print("rsp")
                        writer = pandas.ExcelWriter(FinalFilename)
                    
                        df.to_excel(writer,'Sheet1', index=False)
                        
                        try:
                            writer.save()
                            FileSaved=True
                        except:
                            print("close excel file and try again")
                            rsp=self.DCloseDocToContinue(DocumentNameUsed=FinalFilename)
                            print(rsp)
                            print("rsp")
                            
                    print(FileSaved)
                    print("end of loop FileSaved")
                    #Skip= self.OpenDeleteRecreateSheet(filename=FinalFilename, frame=df,columnList=itemDescription, SqlDatase=False)
                    
                    if rsp==1:
                        self.frameStyleByColor(path=FinalFilename, columnsList=itemDescription, color='blue')
                        os.startfile(FinalFilename)
                        
                        path=os.path.realpath(filePath)
                        os.startfile(path)
                        self.OperaConcluidaWindow()
                    
                    self.ChangeStartupDirectory(Folder='AutoFormFillerKey')
                    
                    
                
        
        
                     
# =============================================================================
#                     
#                     
#                 #print(storedValues.values[rw][0])
#                 #FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
#                 SaVedLocation=storedValues.values[rw][0]
#                 print(SaVedLocation)
#                 print((SaVedLocation.split('.')[0])+".xlsx")
#                 
#                 FinalSaveLocation0=(SaVedLocation.split('.')[0])+".xlsx"
#                 
#                 home = os.path.expanduser('~')
#                 
#                 
#                 filePath = os.path.join(home, 'Documents','AutoFormFillerOutputFiles')
#             
#             
#             
#                 #filePath = r'C:\Users\IgorDC\Documents\AutoFormFillerFiles'
#                 
#                 if not os.path.exists(filePath):
#                     os.makedirs(filePath)
#                 
#                 FinalSaveLocation=filePath +"\\" + FinalSaveLocation0
#                 
#                 
#                 FileSaved=False 
#                 rsp=1
#                 while FileSaved==False and rsp==1:
#                     try:
#                         wb.SaveAs(FinalSaveLocation)
#                         FileSaved=True
#                     except:
#                         rsp=self.DCloseDocToContinue(DocumentNameUsed=FinalSaveLocation)
#                         FileSaved=False
# =============================================================================
        #wb.Close(True)
        
# =============================================================================
#         #path="C:/Users"
#         path=os.path.realpath(filePath)
#         os.startfile(path)
# =============================================================================
        
        
        
        
        
        
        
        
        #for file in filesInFolder:
            


        
        
        
        pass
        
    def FuncLoadToMultipleFiles(self):
        self.OptionForItemInDatabase(option="LoadToMulFiles")
        
        pass
    
    def GetExcelFilesinLocation(self,location, IncludeFolder=False):
        home = os.path.expanduser('~') 
        
        SelectedFolder = os.path.join(home, 'Documents',location)
        
        print(SelectedFolder)
        
# =============================================================================
#         filesInFolder = [f.split("\\")[-1] for f in glob.glob(SelectedFolder + "/*.xl*", recursive=True)]
#         print(filesInFolder)
# =============================================================================
        print("files in folder with ~$")
        
        if IncludeFolder==False:
            filesInFolder = [f.split("\\")[-1] for f in glob.glob(SelectedFolder + "/[!~$]*.xl*", recursive=True)]
            
            print(filesInFolder)
            print("files in folder withOUT ~$")
        if IncludeFolder==True:
            filesInFolder=glob.glob(SelectedFolder+"\\*.xl*")
        
        return filesInFolder


    def valueInList(self,firstList,listTocheck, InListFlag=True, KeepOriginal=True):
        
        print(listTocheck)
        print("listTocheck before before")
            
        firstList1=self.getLocationValues(dataframe=firstList)
        
        listTocheck1=self.getLocationValues(dataframe=listTocheck)
        
        print(firstList1)
        
        print("firstList1 in valueInList1 before")
        
        print(listTocheck1)
        
        print("listTocheck1 in valueInList before")
        
        
        ListofInlist=[]
        ExistInside=False
        for value in range(len(listTocheck1)):
            if InListFlag==True:
                Inlist=listTocheck1[value] in firstList1
            if InListFlag==False:    
                Inlist=listTocheck1[value] not in firstList1
            if Inlist:
                ExistInside=True
                if KeepOriginal==True:
                    ListofInlist.append(listTocheck[value])
                if KeepOriginal==False:
                    ListofInlist.append(listTocheck1[value])
                    
        print(ListofInlist)
        return ExistInside, ListofInlist 
    
    def removeRowsInPandas(self,dataframe,column,listUsed):
        for values in listUsed:
            dataframe=dataframe[~(dataframe[column]==values)]
        
        return dataframe
        
    
    def LoadToMultipleFilesDetails(self, datafillName, FilesInDatabaseLocation, columnsFileinDatabase):
        
        filename="OneFileToMultipleFiles.xlsx"
        
        location=filename
        
        home = os.path.expanduser('~')       
        filePath = os.path.join(home, 'Documents',location)
        
        #filename2=filePath+filename
        
        filesInFolder1=self.GetExcelFilesinLocation(location="AutoFormFillerFiles")
        
        
        filesInFolder2=self.GetExcelFilesinLocation(location="AutoFormFillerOutputFiles")
        
        listTocheck=[filename]+filesInFolder1+filesInFolder2
        print(listTocheck)
        
        
        #listTocheck=self.getLocationValues(dataframe=listTocheck)
        
        
        
        location='AutoFormFillerOutputFiles'
        
        home = os.path.expanduser('~')       
        filePath = os.path.join(home, 'Documents',location)
        
# =============================================================================
#         InFileLocation, SenteceOpenFiles=self.CheckIfInFileLocation(fileToCheck=filename, location=location)
#             
#         if InFileLocation:
#             self.DProriFileNamesUsed(filenameUsed)
# =============================================================================
            
        
        databaseUsed0=self.readSqlDatabase(table_name="KEY_"+datafillName)
        databaseUsed=databaseUsed0[0]
        
        OriginalFileNameTable0=self.readSqlDatabase(table_name=self.FilesInDatabaseLocation)
        print(OriginalFileNameTable0)
        
        OriginalFileNameTable=OriginalFileNameTable0[0]
        print(OriginalFileNameTable)
        print("OriginalFileNameTable")
        OriginalFileNameTable=OriginalFileNameTable[OriginalFileNameTable[self.columnsFileinDatabase[0]]==datafillName]
        
        
        OriFileName0=OriginalFileNameTable[self.columnsFileinDatabase[2]]
        print(OriFileName0)
        OriFileName=OriFileName0.values[0]
        print(OriFileName)
        print("OriFileName after")
        
        print(databaseUsed)
        itemAddress=databaseUsed.columns.values[1:]
        print(itemAddress)
        print("item address above")
        
        itemDescription=databaseUsed.values[0][1:]
        print(itemDescription)
        print("item Description above")
        itemDescription=["Save Name"]+list(itemDescription)
        FrameUsed=pandas.DataFrame(columns=itemDescription)
        
        Skip= self.OpenDeleteRecreateSheet(filename=filename, frame=FrameUsed,columnList=itemDescription, SqlDatase=False)
        if Skip==False:
            self.frameStyleByColor(path=filename, columnsList=itemDescription, color='blue')
            os.startfile(filename)
            rsp=self.AddInfoEXCEL()
            if rsp==QtWidgets.QDialog.Accepted:
                storedValues=pandas.read_excel(filename)

                storedValues0=self.getLocationValues(dataframe=storedValues, column="Save Name")
 
                rsp, storedValues = self.FileExistInDatabaseFunction(dataframe=storedValues0,column="Save Name",filename=filename, listTocheck=listTocheck)        
                
                print(storedValues)
                print("storedValues after")
            
                    
            if rsp==1:
                
                wb = self.openWorkbook(OriFileName)
                wb.Close(True)
                wb = self.openWorkbook(OriFileName)
                self.placeValuesInFile(workbook=wb,itemAddress=itemAddress, storedValues=storedValues, filePath=filePath)
    
        pass
    
    def getLocationValues(self, dataframe, column=None):
        Skip=False
        IsItAList= self.isList(item=dataframe)
    
        if IsItAList:
            PandaslistT = pandas.Series(dataframe)
            print(PandaslistT)
            print("PandaslistT")
            
            Pandaslist=PandaslistT.str.split('\\',expand=True).iloc[:,-1]
            print(Pandaslist)
            print("Pandaslist")
        else:
            try:
                Pandaslist=dataframe[column].str.split('\\',expand=True).iloc[:,-1]
            except IndexError:
                dataframe=pandas.DataFrame(columns=[])
                Skip=True
        
        if Skip==False:
            print(Pandaslist)
            Pandaslist0=Pandaslist.str.split('/',expand=True).iloc[:,-1]
            Pandaslist1=Pandaslist0.str.split('.',expand=True)[0]
            print(Pandaslist)
            Pandaslist2=Pandaslist1+".xlsx"
            print(Pandaslist2)
            
            if IsItAList:
                #dataframe to list
                dataframe=Pandaslist2.tolist() 
            else:
                dataframe[column]=Pandaslist2
                
        return dataframe
    
    def valueInFrame(self,dataframe,column, listTocheck):
        print(dataframe[column])
        ListofInlist=[]
        ExistInside=False
        for value in listTocheck:
            print(value)
            print(dataframe[column])
            Inlist=value in list(dataframe[column])
            if Inlist:
                ExistInside=True
                ListofInlist.append(value)
        print(ListofInlist)
        return ExistInside, ListofInlist
    
    def FileExistInDatabaseFunction(self,dataframe,column,filename,listTocheck):
        rsp=1
        ExistInside=True
        self.IgonoreClicked=False
        ExistsNan=True
        rsp2=1
        SkipExistInside=True
        while rsp==1 and rsp2==1 and (ExistInside==True or ExistsNan==True):
           
            dataframe=pandas.read_excel(filename)
            
            dataframe=self.getLocationValues(dataframe=dataframe, column="Save Name")
            
            ExistsNan=self.CheckAllFilled(dataframe, anyNan=False)
            if ExistsNan:
                rsp2=self.AddInfoEXCEL()
            try:
                
                ExistInside, ListofInlist0 = self.valueInFrame(dataframe=dataframe,column=column, listTocheck=listTocheck)
                SkipExistInside=False
            except KeyError:
                pass
            
            if SkipExistInside==False:
                print(ExistInside)
                print("ExistInside")
                ListofInlist=self.ListToSentence(ListofInlist0)
                if ExistInside==True:
                   rsp=self.DProriFileNamesUsed(ListofInlist) 
              
            print(ExistsNan)
            print("ExistsNan")
# =============================================================================
#         if rsp==0:
#             print("rsp =0")
# =============================================================================
        if self.IgonoreClicked:
            print("button clicked")
            for row in range(len(ListofInlist0)):
                dataframe=dataframe[~(dataframe[column]==ListofInlist0[row])]
            print(dataframe)
            rsp=1
            
        if rsp2==0:
            rsp=0
            
        return rsp, dataframe
    
    def FuncIgnore(self):
        self.IgonoreClicked=True
      
        print(self.IgonoreClicked)
    
    
    def getAddressesFromColumn(self, itemAddress):
        
        rowList=[]
        columnList=[]
        for values in range(len(itemAddress)):
            temp = re.findall(r'\d+', itemAddress[values]) 
            i,c = list(map(int, temp)) 
            rowList.append(i)
            columnList.append(c)
            print(i)
            print(c)
        print(columnList)
        print("columnList")
        print(rowList)
        print("rowList")
        
        return rowList, columnList
    
    
    def placeValuesInFile(self, workbook,itemAddress, storedValues, filePath):
        wb = workbook
        
        print(os.getcwd())
        
        documentSaved=False
        rsp=1
        while documentSaved==False and rsp==1:
            
            try:
                
                ws = wb.Sheets(1)
                documentSaved=True
            except:
                rsp=self.addInformationANDSave()
                print("exception runned")

        if rsp==1:    
            allData = ws.UsedRange
            
            excel = win32.gencache.EnsureDispatch('Excel.Application') 
            
            excel.Visible = True
            
            # Get number of rows used on active sheet
            ind = allData.Rows.Count
            print ('Number of rows used in sheet : ', ind)
            
            #Get number of columns used on active sheet
            col = allData.Columns.Count
            print ('Number of columns used in sheet : ', col)
            
            print(col)
            
            print(ind)
            
# =============================================================================
#             rowList=[]
#             columnList=[]
#             for values in range(len(itemAddress)):
#                 temp = re.findall(r'\d+', itemAddress[values]) 
#                 i,c = list(map(int, temp)) 
#                 rowList.append(i)
#                 columnList.append(c)
#                 print(i)
#                 print(c)
#             print(columnList)
#             print("columnList")
#             print(rowList)
#             print("rowList")
# =============================================================================
            
            rowList, columnList=self.getAddressesFromColumn(itemAddress=itemAddress)
            
# =============================================================================
#             print(storedValues)
#             print(storedValues.values[0])
#             print(storedValues.values[1])
#             print(storedValues.values[2])
#             print(storedValues.values[3])
# =============================================================================
            
            numberOfRows=len(storedValues.index)
            print(numberOfRows)
            
            #rw=0
            
            #if 1==True:
                
            for rw in range(numberOfRows):
                for cl in range(1,len(itemAddress)+1):
                    print(storedValues.values[rw][cl])
                    #print(type(storedValues.values[rw][cl]))
                    print(str(storedValues.values[rw][cl]).upper()=="NAN")
                    i=rowList[cl-1]
                    c=columnList[cl-1]
                    if str(storedValues.values[rw][cl]).upper()!="NAN":
                        print("ok")
                        ws.Cells(i,c).Value=storedValues.values[rw][cl]
                    else:
                        print("Not Nan ok")
                        ws.Cells(i,c).Value=''
                #print(storedValues.values[rw][0])
                #FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
                SaVedLocation=storedValues.values[rw][0]
                print(SaVedLocation)
                print((SaVedLocation.split('.')[0])+".xlsx")
                
                FinalSaveLocation2=(SaVedLocation.split("\\")[-1])
                
                print(FinalSaveLocation2)
                
                
                FinalSaveLocation1=(FinalSaveLocation2.split("/")[-1])
                
                print(FinalSaveLocation1)
                
                FinalSaveLocation0=(FinalSaveLocation1.split('.')[0])+".xlsx"
                
                print(FinalSaveLocation0)
                
# =============================================================================
#                 home = os.path.expanduser('~')
#                 
#                 
#                 filePath = os.path.join(home, 'Documents','AutoFormFillerOutputFiles')
# =============================================================================
            
            
            
                #filePath = r'C:\Users\IgorDC\Documents\AutoFormFillerFiles'
                
                if not os.path.exists(filePath):
                    os.makedirs(filePath)
                
                FinalSaveLocation=filePath +"\\" + FinalSaveLocation0
                
                
                print(FinalSaveLocation)
                print("FinalSaveLocation")
                
                
                FileSaved=False 
                rsp=1
                while FileSaved==False and rsp==1:
                    try:
                        wb.SaveAs(FinalSaveLocation)
                        FileSaved=True
# =============================================================================
                    except Exception as e:
                        print(e)
# =============================================================================
#                         xl = Dispatch('Excel.Application')
#                         wb = xl.Workbooks.Open(FinalSaveLocation)
#                         wb.Close(True)
#                         rsp=1
# =============================================================================
                        print("go again")
                        rsp=self.DCloseDocToContinue(DocumentNameUsed=FinalSaveLocation)
                        FileSaved=False
# =============================================================================
            wb.Close(True)
            
            #path="C:/Users"
            path=os.path.realpath(filePath)
            os.startfile(path)
            self.OperaConcluidaWindow()
            
            
# =============================================================================
#             listOfValues=[]
#             listOfValueNames=[]
#             
#             for i in range(1,ind+1):
#                 for c in range(1,col+1):
# =============================================================================
                    
                    #ws.Cells(i,c).Value=value
           
# =============================================================================
#             for i in range(1,ind+1):
#                 
#                 for c in range(1,col+1):
#                     
#                     value = ws.Cells(i,c).Value
#                     
#                     if value == "XYXYXYXYX":
#                         listOfValueNames.append("["+ str(i) +","+ str(c)+"]")
#                         listOfValues.append([i,c])
# 
#             NumList= len(listOfValues)
#             return NumList, listOfValueNames, rsp
# =============================================================================
        
        if rsp==0:
            NumList=0
            listOfValueNames= []
            
            return NumList, listOfValueNames, rsp
        
        
        pass
    
    def dropSqlTable(self, database,table):
        connection= sqlite3.connect(database)
        cursor= connection.cursor()
        print(table)
        dropTableStatement= f"""DROP TABLE '{table}'"""
        try:
            cursor.execute(dropTableStatement)
            print("table dropped")
        except:
            print("table does not exist")
            pass
            
        connection.close()
        
        
    def uploadDatabase(self, SqlDatase):
        FilesInDatabaseLocation=self.FilesInDatabaseLocation
        columnsFileinDatabase=self.columnsFileinDatabase
        
        if SqlDatase==True:
            FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
        
        FileExists=False
        
        if SqlDatase==False:
            try:
                FilesInDatabase=pandas.read_excel(FilesInDatabaseLocation)
                FileExists=True
                
            except FileNotFoundError:
                FileExists=False
                #self.ui0.tableWidget.clearContents()
    # =============================================================================
    #             self.ui0.tableWidget.setRowCount(0)
    #             self.ui0.tableWidget.setRowCount(7)
    #             
    #             header=self.ui0.tableWidget.horizontalHeader()
    #             header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
    #             header.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
    #             header.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
    #             self.ui0.tableWidget.horizontalHeader().setDefaultSectionSize(132)
    # =============================================================================
                
            pass
        
        if SqlDatase==True:
            FilesInDatabase , FileExists = self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
            
        
        if FileExists==False:
            self.ui0.tableWidget.setRowCount(0)
            self.ui0.tableWidget.setRowCount(7)
            
            header=self.ui0.tableWidget.horizontalHeader()
            header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            header.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            header.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            self.ui0.tableWidget.horizontalHeader().setDefaultSectionSize(132)
        
        if FileExists==True:
# =============================================================================
#             FilesInDatabase["UPPER"]=FilesInDatabase[columnsFileinDatabase[0]].str.upper()
#             
#             pandas.DataFrame.sort_values(FilesInDatabase, by=["UPPER", columnsFileinDatabase[0],columnsFileinDatabase[1],columnsFileinDatabase[2]], inplace=True)
#             FilesInDatabase=FilesInDatabase[columnsFileinDatabase]
# =============================================================================
            
            
            FilesInDatabase=FilesInDatabase[columnsFileinDatabase]
            
            NumbRowsPandas=len(FilesInDatabase.index)
            print(len(FilesInDatabase.index))
            print("len(FilesInDatabase.index)")
            
            if NumbRowsPandas ==0:
                #self.ui0.tableWidget.clearContents()
                # self.ui0.tableWidget.setRowCount(0)
                # self.ui0.tableWidget.setRowCount(7)
                
                # header=self.ui0.tableWidget.horizontalHeader()
                # header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
                # header.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
                # header.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
                # self.ui0.tableWidget.horizontalHeader().setDefaultSectionSize(132)
                pass

            else:
                print(FilesInDatabase)
                
                # #self.ui.tableWidget.setRowCount(0)
                # self.ui0.tableWidget.setRowCount(0)
                # #self.ui.tableWidget.setRowCount(0)
                
                # for rowNumber, rowData in enumerate(FilesInDatabase.values):
                #     self.ui0.tableWidget.insertRow(rowNumber)
                #     for columnNumber , data in enumerate(rowData):
                #         #self.ui.tableWidget.insertColumn(columnNumber)
                #         self.ui0.tableWidget.setItem(rowNumber, columnNumber, QtWidgets.QTableWidgetItem(str(data)))
                
                # header=self.ui0.tableWidget.horizontalHeader()
                # header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
                # header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
                # header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
                
        pass
    
    
    def CheckIfInFileLocation(self,fileToCheck, internalLocation='Documents', location='AutoFormFillerFiles'):
        

        home = os.path.expanduser('~')
             
        print(home)
         
        filePath = os.path.join(home, internalLocation ,location)

        
        if not os.path.exists(filePath):
            os.makedirs(filePath)
         
        files = [f.split("\\")[-1] for f in glob.glob(filePath + "/*", recursive=True)]
        
        print(files)
        
        #if multipleFiles==False:
        
        if self.isList(item=fileToCheck):
            if len(fileToCheck)==1:
                if fileToCheck[0] in files:
                    InFileLocation=True
                    SenteceOpenFiles=str(fileToCheck)
                else:
                    InFileLocation=False
                    SenteceOpenFiles=''
                    
            #if multipleFiles==True:
            if self.isList(item=fileToCheck):    
                if len(fileToCheck)>1:
                    InFileLocation0=[]
                    ListOfOpenFiles=[]
                    for single in fileToCheck:
                        if single in files:
                            
                            value=True
                            InFileLocation0.append(value)
                            ListOfOpenFiles.append(single)
                    print(InFileLocation0)
                    InFileLocation=True in InFileLocation0
                    SenteceOpenFiles0=self.ListToSentence(ListOfOpenFiles)
                    SenteceOpenFiles=str(SenteceOpenFiles0)
                    if InFileLocation==False:
                        SenteceOpenFiles=''
                        
        if self.isList(item=fileToCheck)==False:
            if fileToCheck in files:
                    InFileLocation=True
                    SenteceOpenFiles=str(fileToCheck)
            else:
                InFileLocation=False
                SenteceOpenFiles=''
                
        return InFileLocation, SenteceOpenFiles
    
    
    def ListToSentence(self,ListUsed):
        listb=str(ListUsed)
        listc=listb[1:-1]
        return listc
    
    def isList(self,item):
        IsthisList=type(item)==type([])
        return IsthisList  
    
    def funcStartswithTemp(self,filename):
    
        checkstring=filename.split("/")[-1]
    
        StartswithTemp=checkstring.startswith('temp_')
    
        return StartswithTemp
        
    def clickedGetFileName(self):
        
        UseExcel=self.UseExcel
        SqlDatase=self.SqlDatase
        
        
        
        updateNamesFile='updateNamesFile.xlsx'
        FilesInDatabaseLocation=self.FilesInDatabaseLocation
        
        if SqlDatase==True:
            updateNamesFile=updateNamesFile.split('.')[0]
            FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
        
        columnsFileinDatabase=self.columnsFileinDatabase
        
        
        
        
        
        
        internalLocation='Documents'
        excelLocation='AutoFormFillerFiles'
        
        fname=''
        print("ok add file clicked")
        
        home = os.path.expanduser('~')
        
        print(home)
        
        filePath = os.path.join(home, internalLocation ,excelLocation)
        
        location = os.path.join(home, internalLocation)
        
        StartswithTemp=True
        
        rsp=1
        
        while StartswithTemp==True and rsp==1:
            
            fname, _ = QFileDialog.getOpenFileName(self, 'Open a file to add to the database',location , "Excel Files (*.xl* *.xlsx *.xlsm *.xlsb *.xlam *.xltx *.xltm *.xls *.xla *.xlt *.xlm *.xlw)")
            
            print(fname)
            
            StartswithTemp=self.funcStartswithTemp(filename=fname)
            
            if StartswithTemp:
                rsp=self.FileCanNotStartWith()
                if rsp==0:
                    fname=''
        
        #fname = QFileDialog.getOpenFileName(self, 'Open file','/home')
        
        if fname !='':
            
            #datafillName, FilesInDatabaseExists=self.CheckJobNameRedo(SqlDatase=SqlDatase, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
        
            print(fname)
            
            print("fname")
            
            fileNameOnly=fname.split("/")[-1]
            print(fileNameOnly)
            
# =============================================================================
#             home = os.path.expanduser('~')
#             
#             print(home)
# =============================================================================
            
            filePath = os.path.join(home, 'Documents','AutoFormFillerFiles')
            InFileLocation, SenteceOpenFiles=self.CheckIfInFileLocation(fileToCheck=fileNameOnly, location=excelLocation)
            
# =============================================================================
#             
#             
#             
#             #filePath = r'C:\Users\IgorDC\Documents\AutoFormFillerFiles'
#             
#             if not os.path.exists(filePath):
#                 os.makedirs(filePath)
#              
#             files = [f.split("\\")[-1] for f in glob.glob(filePath + "/*", recursive=True)]
#             
#             print(files)
#             
#             if fileNameOnly in files:
# =============================================================================
            if InFileLocation==True:
                print("there is the same file inside")
                
                rsp=self.fileAlreadyExits()
                
                if rsp==1:
                    
# =============================================================================
#                     datafillName, FilesInDatabaseExists, rsp2=self.CheckJobNameRedo(SqlDatase=SqlDatase, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
#                     
#                     if rsp2==1:
# =============================================================================
                    
                    newFileLocation0=filePath +"\\" + fileNameOnly
                    print(newFileLocation0)
                
                    print("newFileLocation not coppied")
                
                    newFileLocation=filePath +"\\temp_" + fileNameOnly
                    print(newFileLocation)
                    
                    print("newFileLocation")
                    copiedDoc=False
                    rsp=1
                    while copiedDoc==False and rsp==1:
                        try:
                            shutil.copy2(fname, newFileLocation)
                            copiedDoc=True
                        except PermissionError:
                            rsp=self.InicioSalveDocumento()
                        
                else:
                    fname=''
                
            else:
# =============================================================================
#                 datafillName, FilesInDatabaseExists, rsp2=self.CheckJobNameRedo(SqlDatase=SqlDatase, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
#                     
#                 if rsp2==1:
# =============================================================================
                print("add file to this folder")
                
                newFileLocation0=filePath +"\\" + fileNameOnly
                print(newFileLocation0)
                
                print("newFileLocation")
                
                shutil.copy2(fname, newFileLocation0)
                
                newFileLocation=filePath +"\\temp_" + fileNameOnly
                print(newFileLocation)
                
                print("newFileLocation")
                copiedDoc=False
                rsp=1
                while copiedDoc==False and rsp==1:
                    try:
                        shutil.copy2(fname, newFileLocation)
                        copiedDoc=True
                    except PermissionError:
                        rsp=self.InicioSalveDocumento()
                
            if rsp==1:
                if fname=='':
                    print("cancel button clicked")
                
                else:
                    datafillName, FilesInDatabaseExists, rsp2=self.CheckJobNameRedo(SqlDatase=SqlDatase, FilesInDatabaseLocation=FilesInDatabaseLocation, columnsFileinDatabase=columnsFileinDatabase)
                    try:
                        if self.IgonoreClicked: 
                            rsp2=1
                    except:
                        pass
                    
                    if rsp2==1:
                    
                        rsp=0
                        
                        #os.startfile(fname)
                        print("before openWorkbook")
                        #fname=fname.replace('/', '\\')
                        #fname="%r"%fname
                        #print(fname)
                        #wb = self.openWorkbook(r"C:\\Users\\IgorDC\\Desktop\\Notebook\\ExcelTestFile.xlsx")
                        #wb = self.openWorkbook(fname)
                        
                        #### Note the function self.openWorkbook does not need to change when using sql database
                        #### for it is used to open the excel file used to mark all input locations
                        wb = self.openWorkbook(newFileLocation)
                        
                        print("after openWorkbook")
                    
                        if wb!=None:
                        
                            #rsp=self.CopyPasteWindow()
                            
            # =============================================================================
            #                 if rsp==1:
            #                     NumList=0
            # =============================================================================
                            NumList=0
                            rsp=1
                            
                            while NumList==0 and rsp==1:
                                
                                rsp=self.CopyPasteWindow()
                                
                                #NumList, listOfValueNames, rsp =self.checkListValues(excelFileLocation=fname, workbook=wb, rsp=rsp)
                                NumList, listOfValueNames, rsp =self.checkListValues(excelFileLocation=newFileLocation, workbook=wb, rsp=rsp)
                                
                                print(NumList)
                                
                            if rsp==1:
                                
                                print("list has more then one value")
                                
                                self.clickedAddFile(listOfValueNames=listOfValueNames, newFileSavedLocation=newFileLocation0, UseExcel=UseExcel, SqlDatase=SqlDatase, datafillName=datafillName,FilesInDatabaseExists=FilesInDatabaseExists)
                                
                               
                                
                                
                            
                            
                # =============================================================================
                #             cb = QtWidgets.QApplication.clipboard()
                #             cb.clear(mode=cb.Clipboard )
                #             cb.setText("XYXYXYXYX", mode=cb.Clipboard)
                # =============================================================================
                        
                    pass
        
    
    def openWorkbook(self, xlfile):
        rsp=1
        noerror=False
        xlwb = None    
        while rsp==1 and noerror==False:
            try:
                
                xlapp = win32.gencache.EnsureDispatch('Excel.Application')
                
# =============================================================================
#                 xlapp.ScreenUpdating = False
#                 xlapp.DisplayAlerts = False
#                 xlapp.EnableEvents = False
# =============================================================================
                xlapp.DisplayAlerts = False
                print("no eror -1")
                xlapp.Visible = True
                print("no eror 0")
                noerror=True
                rsp=1
                print("no eror")
            except TypeError:
                print('sheet is open and has changes')
                rsp = self.InicioSalveDocumento()
                #noerror=False
        
        if rsp==1:
            
            if noerror==True:
        
                    try:
                        print("no eror 1")
                        xlapp = win32.gencache.EnsureDispatch('Excel.Application')
# =============================================================================
#                         xlapp.ScreenUpdating = False
#                         xlapp.DisplayAlerts = False
#                         xlapp.EnableEvents = False
# =============================================================================
                        
                        xlapp.DisplayAlerts = False
                        xlapp.Visible = True
                        print("no eror 2")
                        xlwb = xlapp.Workbooks.Open(xlfile)
                        #xlapp.Visible = True
                        print("no eror 3")
                        #xlapp.DisplayAlerts = False
                    except Exception as e:
                        print("no eror exception")
                        print(e)
                        xlapp.DisplayAlerts = False
                        map(lambda book: book.Close(False), xlapp.Workbooks)
                        xlapp.Quit()
                        
                        xlapp = win32.gencache.EnsureDispatch('Excel.Application') 
# =============================================================================
#                         xlapp.ScreenUpdating = False
#                         xlapp.DisplayAlerts = False
#                         xlapp.EnableEvents = False
# =============================================================================
                        
                        
                        xlapp.Visible = True
                        xlapp.DisplayAlerts = False
                        
                        xlwb = xlapp.Workbooks.Open(xlfile)
                        #xlapp.Visible = True
                        
                        
# =============================================================================
#         xlapp.ScreenUpdating = True
#         xlapp.DisplayAlerts = True
#         xlapp.EnableEvents = True        
# =============================================================================
        return(xlwb)
    
    
    
    
    def checkListValues(self,excelFileLocation, workbook, rsp):
        
        wb = workbook
        
        print(os.getcwd())
        
        documentSaved=False
        
        while documentSaved==False and rsp==1:
            
            try:
                
                ws = wb.Sheets(1)
                documentSaved=True
            except:
                rsp=self.addInformationANDSave()
                print("exception runned")

        if rsp==1:    
            allData = ws.UsedRange
            
            excel = win32.gencache.EnsureDispatch('Excel.Application') 
            
            excel.Visible = True
            
            # Get number of rows used on active sheet
            ind = allData.Rows.Count
            print ('Number of rows used in sheet : ', ind)
            
            #Get number of columns used on active sheet
            col = allData.Columns.Count
            print ('Number of columns used in sheet : ', col)
            
            print(col)
            
            print(ind)
            
            listOfValues=[]
            listOfValueNames=[]
           
            for i in range(1,ind+1):
                
                for c in range(1,col+1):
                    
                    value = ws.Cells(i,c).Value
                    
                    if value == "XYXYXYXYX":
                        listOfValueNames.append("["+ str(i) +","+ str(c)+"]")
                        listOfValues.append([i,c])

            NumList= len(listOfValues)
            return NumList, listOfValueNames, rsp
        
        if rsp==0:
            NumList=0
            listOfValueNames= []
            
            return NumList, listOfValueNames, rsp
        
    def CheckJobNameRedo(self, SqlDatase, FilesInDatabaseLocation, columnsFileinDatabase):
        Redowhile=True
        rsp=1
        datafillName=''
        
        print("Job name Window Started Now")
        while (Redowhile==True or datafillName =='')  and rsp==1:
            rsp, datafillName=self.NewJobNameWindow()
# =============================================================================
#             if rsp==1:
#                 rsp=self.DRewriteDatafillJobName(filenameUsed=datafillName)
# =============================================================================
            
            datafillName=datafillName.split()
            datafillName=" ".join(datafillName)
            
            if SqlDatase==False:
                ####Added part below
                FilesInDatabaseExists=False
                try:
                    FilesInDatabase=pandas.read_excel(FilesInDatabaseLocation)
                    FilesInDatabaseExists=True
                except FileNotFoundError:
                    print('database DOES NOT exist... continue')
                    FilesInDatabaseExists=False
                    
            if SqlDatase==True:
                FilesInDatabase , FilesInDatabaseExists = self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
                    
            if FilesInDatabaseExists==True:
                
                if datafillName in list(FilesInDatabase[columnsFileinDatabase[0]]):
                    Redowhile=True
                    if rsp==1:
                        rsp=self.DRewriteDatafillJobName(filenameUsed=datafillName)
                else:
                   Redowhile=False 
            
            if FilesInDatabaseExists==False:
                Redowhile=False
                
        return datafillName, FilesInDatabaseExists, rsp
        
    
    def clickedAddFile(self,listOfValueNames, newFileSavedLocation, SqlDatase, datafillName, FilesInDatabaseExists, UseExcel=True):
        Skip=False
        
        updateNamesFile='updateNamesFile.xlsx'
        FilesInDatabaseLocation=self.FilesInDatabaseLocation
        
        if SqlDatase==True:
            updateNamesFile=updateNamesFile.split('.')[0]
            FilesInDatabaseLocation=FilesInDatabaseLocation.split('.')[0]
        
        columnsFileinDatabase=self.columnsFileinDatabase
        
# =============================================================================
#         datafillName=''
#         SheetName=''
#         newFileSavedLocation=''
#         InputData=[datafillName,SheetName,newFileSavedLocation]
# =============================================================================
        
        
        ##### TODO add a window for itens without needing to write all code (from function instead)
        #self.JobNameWindow()
        
# =============================================================================
#         Redowhile=True
#         rsp=1
#         datafillName=''
#         while (Redowhile==True or datafillName =='')  and rsp==1:
#             rsp, datafillName=self.JobNameWindow()
#             
#             datafillName=datafillName.split()
#             datafillName=" ".join(datafillName)
#             
#             if SqlDatase==False:
#                 ####Added part below
#                 FilesInDatabaseExists=False
#                 try:
#                     FilesInDatabase=pandas.read_excel(FilesInDatabaseLocation)
#                     FilesInDatabaseExists=True
#                 except FileNotFoundError:
#                     print('database DOES NOT exist... continue')
#                     FilesInDatabaseExists=False
#                     
#             if SqlDatase==True:
#                 FilesInDatabase , FilesInDatabaseExists = self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
#                     
#             if FilesInDatabaseExists==True:
#                 
#                 if datafillName in list(FilesInDatabase[columnsFileinDatabase[0]]):
#                     Redowhile=True
#                 else:
#                    Redowhile=False 
#             
#             if FilesInDatabaseExists==False:
#                 Redowhile=False
# =============================================================================
        
        ####Added part above
        
        #if rsp==1 and datafillName !='':
            
        ExistsNan=True
        rsp=1
        
        if UseExcel==True:
            UpdateValues=pandas.DataFrame(columns=listOfValueNames)
            Skip= self.OpenDeleteRecreateSheet(filename=updateNamesFile, frame=UpdateValues,columnList=listOfValueNames, SqlDatase=SqlDatase)
            if Skip==False:
                self.frameStyleByColor(path=updateNamesFile, columnsList=listOfValueNames, color='blue')
                os.startfile(updateNamesFile)
        if UseExcel==False:
            Skip=False
            sizeFrame=len(listOfValueNames)
            frame=['']*sizeFrame
            
        while rsp==1 and ExistsNan==True and Skip==False:
            print(rsp)
            print("rsp after while loop")
            print(ExistsNan)
            print("ExistisNan after while loop")
            print(Skip)
            print("Skip after while loop")
            if UseExcel==True:
                rsp=self.AddInfoEXCEL()
                storedValues=pandas.read_excel(updateNamesFile)
                ExistsNan=self.CheckAllFilled(storedValues=storedValues)
                
            if UseExcel==False:
                rsp, frame=self.addInfoTable(frame=frame, columnNames=listOfValueNames)
                ExistsNan=self.Inframe(frame)

        print(rsp)
        print("rsp after while loop end")
        print(ExistsNan)
        print("ExistisNan after while loop end")
        print(Skip)
        print("Skip after while loop end")

                        
        if rsp==QtWidgets.QDialog.Accepted:
            print("Ok now it starts updating data")
            
            SheetName="KEY_" + datafillName.split('.')[0] + '.xlsx'    

            if SqlDatase==True:
                SheetName=SheetName.split('.')[0]
            
            print(SheetName)
            if UseExcel==True:
                storedValues=pandas.read_excel(updateNamesFile)
                Skip= self.OpenDeleteRecreateSheet(filename=SheetName, frame=storedValues,columnList=listOfValueNames, SqlDatase=SqlDatase)
                print("before open file 222222 UseExcel==True")
                
            if UseExcel==False:
                #### TODO create excel file for sheet with only frame names
                print(listOfValueNames)
                print("the above is the list of columns")
                
                storedValues=pandas.DataFrame(columns=listOfValueNames)
                
                print(frame)
                print('frame before for loop')
                for i in range(len(listOfValueNames)):
                    storedValues[listOfValueNames[i]]=[frame[i]]
                
                print(storedValues)
                
                print("this is the endframe just created")
                Skip= self.OpenDeleteRecreateSheet(filename=SheetName, frame=storedValues,columnList=listOfValueNames, SqlDatase=SqlDatase)
            
            if Skip==False:
                
                InputData=[datafillName,SheetName,newFileSavedLocation]
                
                
                #self.frameStyleByColor(path=updateNamesFile, columnsList=listOfValueNames, color='blue')
                
                #os.startfile(SheetName)
                
                if SqlDatase==False:
                    FilesInDatabaseExists=False
                    try:
                        FilesInDatabase=pandas.read_excel(FilesInDatabaseLocation)
                        FilesInDatabaseExists=True
                    except FileNotFoundError:
                        print('database DOES NOT exist')
                        FilesInDatabaseExists=False
# =============================================================================
#                         FilesInDatabase=pandas.DataFrame(columns=[])
#                         for i in range(len(columnsFileinDatabase)):
#                             #print(FilesInDatabase[columnsFileinDatabase[i]])
#                             FilesInDatabase[columnsFileinDatabase[i]]= pandas.Series(InputData[i])
# =============================================================================
                
                if SqlDatase==True:
                    FilesInDatabase , FilesInDatabaseExists = self.readSqlDatabase(table_name=FilesInDatabaseLocation,columns=columnsFileinDatabase)
                    
                if FilesInDatabaseExists==False:
                    FilesInDatabase=pandas.DataFrame(columns=[])
                    for i in range(len(columnsFileinDatabase)):
                        #print(FilesInDatabase[columnsFileinDatabase[i]])
                        FilesInDatabase[columnsFileinDatabase[i]]= pandas.Series(InputData[i])
                
                if FilesInDatabaseExists==True:
                    if self.IgonoreClicked:
                        FilesInDatabase=FilesInDatabase[~(FilesInDatabase[columnsFileinDatabase[0]]==datafillName)]
                        self.IgonoreClicked=False
                    
                    
                    print("database exists")
                    TempFilesInDatabase=pandas.DataFrame(columns=[])
                    for i in range(len(columnsFileinDatabase)):
                        #print(TempFilesInDatabase[columnsFileinDatabase[i]])
                        TempFilesInDatabase[columnsFileinDatabase[i]]= pandas.Series(InputData[i])
                        
                    print(TempFilesInDatabase)
                    FilesInDatabase=pandas.concat([FilesInDatabase,TempFilesInDatabase])
                    
                    FilesInDatabase["UPPER"]=FilesInDatabase[columnsFileinDatabase[0]].str.upper()
    
                    pandas.DataFrame.sort_values(FilesInDatabase, by=["UPPER", columnsFileinDatabase[0],columnsFileinDatabase[1],columnsFileinDatabase[2]], inplace=True)
                    #FilesInDatabase=FilesInDatabase[columnsFileinDatabase]
                    

                    print(FilesInDatabase)
                    
                
                
                Skip= self.OpenDeleteRecreateSheet(filename=FilesInDatabaseLocation, frame=FilesInDatabase,columnList=columnsFileinDatabase, SqlDatase=SqlDatase)
                
                if Skip==False:
                    
                    self.uploadDatabase(SqlDatase=SqlDatase)
                    
                    self.OperaConcluidaWindow()
                 
                       
            
                        
            
            
            
            
            
            
        
    
    
    def OpenDeleteRecreateSheet(self,filename,frame,columnList, SqlDatase=True, noColumnList=False):
    
        Skip=False
        if SqlDatase==False:
                    Skip=self.preOpenDeleteRecreateSheet(filename=filename,frame=frame,columnList=columnList, noColumnList=noColumnList)
                    
        if SqlDatase==True:
            self.writeToSqlite(frame=frame, table_name=filename)
            Skip=False
            
        return Skip
    
        
    def preOpenDeleteRecreateSheet(self, filename,frame,columnList, noColumnList):
        
        ### Open delete sheet so it can be saved later below 2
        Skip=False
        rsp=1
        
        writer = pandas.ExcelWriter(filename)
        print('check 111111')
        try:   
            wb=openpyxl.load_workbook(filename)
            print('check 222222')
            sheetloaded=True
        except:
            sheetloaded=False
            ProblemSaving=False
            pass
        
        if sheetloaded==True:    
            print('check 222333')
            del wb['Sheet1']
            print('check 333333')
            
            try:
                writer.save()
                ProblemSaving=False
            except:
                ProblemSaving=True
                
        if ProblemSaving==True:
            #xl = Dispatch('Excel.Application')
            Skip=False
            
            SkipWhileLoop=False
            secondException=False
            while SkipWhileLoop==False and Skip==False:
                print("start loop SkipWhileLoop==False")
                try:
                    xl = Dispatch('Excel.Application')
                    wb = xl.Workbooks(filename)
                    SkipWhileLoop=True
                    Skip=False
                except Exception as e:
                    print(e)
                    f=str(e)
                    print(type(f))
                    print(f)
                    g=''
                    try: 
                        g=e.args[1]
                        if g== 'Exception occurred.':
                            secondException=True
    
                    except Exception as d:
                        print(d)
                        print("exeption of exception")
                        pass
                    
                    if  f=="Excel.Application.Workbooks":
                        rsp=self.InicioSalveDocumento()
    
                    if rsp==QtWidgets.QDialog.Accepted:
                        Skip=False
                        SkipWhileLoop=False
                        if secondException:
                            print("secondException=True runned")
                            SkipWhileLoop=True
                        pass
                    
                    if rsp==0:
                        Skip=True
                        SkipWhileLoop=False
    
            if Skip==False:
                if secondException==False:
                    writer.save()
                    wb.Close(True)
    ##### Open delete sheet so it can be saved later above 2
        
        if Skip==False:
            
            writer = pandas.ExcelWriter(filename)
            
            print(frame)
            print('frame before column filtering')
            
            #### noColumnList=False added so columnlistfilter can be ignored
            if noColumnList==False:
                frame=frame[columnList]
            
            print(frame)
            print("frame with unamed")
    
            frame.drop(frame.columns[frame.columns.str.contains('unnamed',case = False)],axis = 1, inplace = True)
            print("step 1111111" )
            #frame.to_excel(writer,'Sheet1', index=False)
            frame.to_excel(writer, index=False)
            print("step 22222222" )
            try:
              
                writer.save()
                print("step 333333333" )
            except PermissionError:
                xl = Dispatch('Excel.Application')
                print("step 44444444444" )
                wb = xl.Workbooks(filename)
                print("step 55555555" )
                # do some stuff
                wb.Close(True) # save the workbook
                print("step 66666666" )
                writer.save()
                print("step 777777777" )
        
            print(Skip)
            print("skip")
        
        return Skip
        
    
    
    
        
        
        
    def frameStyleByColor(self,path, columnsList, color):
        
        if color=='blue':
            self.frameStyle(path=path, columnsList=columnsList)
            pass
        
        if color=='green':
            self.frameStyle(path=path, columnsList=columnsList, BackgroundRGBr=0 , BackgroundRGBg= 195, BackgroundRGBb=0)
            pass
        
        if color=='red':
            self.frameStyle(path=path, columnsList=columnsList, BackgroundRGBr=237 , BackgroundRGBg= 0, BackgroundRGBb=0)
            pass
        
        if color=='yellow':
            self.frameStyle(path=path, columnsList=columnsList, BackgroundRGBr=255 , BackgroundRGBg= 255, BackgroundRGBb=0, LetterRGBr=0, LetterRGBg=0, LetterRGBb=0)
            pass
        
    
    
    def frameStyle(self,path,columnsList,BackgroundRGBr=0,BackgroundRGBg=0,BackgroundRGBb=206,LetterRGBr=255,LetterRGBg=255,LetterRGBb=255):

        DictionaryofWidths= {x: 1.3*(int(len(x))+13) for x in columnsList}

        frame2 = StyleFrame.read_excel(path= path, sheet_name='Sheet1')
        print(frame2)
        print('ok 1111')
        
        print(columnsList)
        
        print('columnsList used in frameStyle')
         
        ew = StyleFrame.ExcelWriter(path)
        
        print('ok 2222')
         
        HeaderStyle= Styler(bg_color=self.DecHex(BackgroundRGBr)+self.DecHex(BackgroundRGBg)+self.DecHex(BackgroundRGBb), bold=True, font_color= self.DecHex(LetterRGBr)+self.DecHex(LetterRGBg)+self.DecHex(LetterRGBb), border_type='medium', horizontal_alignment='center', vertical_alignment='center', shrink_to_fit=False)
         
        print('ok 3333')
        
        frame2.apply_headers_style(HeaderStyle, style_index_header=True)
         
        print('ok 4444')
        StyleFrame.set_column_width_dict(frame2, DictionaryofWidths)
        
        StyleFrame.to_excel(frame2, excel_writer = ew, sheet_name='Sheet1')

        print('ok 5555')
        
        ew.save()
        
        
    def DecHex(self,n):
        '''
        :para n: int
        :return: str
        >>> DecHex(10)
        'A'
        >>> DecHex(15)
        'F'
        >>> DecHex(32)
        '20'
        >>> DecHex(255)
        'FF'
        >>> DecHex(65535)
        'FFFF'
        '''
    
        x16 = '0 1 2 3 4 5 6 7 8 9 a b c d e f'.upper().split()
        result = []
        try:
            n = int(n)
            if n == 0:
                return '00'
            if 17 > n:
                result.append('0' + x16[(n % 16)])
                n = n // 16
            result.reverse()
            while n > 0:
                result.append(x16[(n % 16)])
                n = n // 16
            result.reverse()
        except ValueError as e:
            return ('Erro: %s' %e)
        except:
            raise
        else:
            return ''.join(result)
        
    
    #def CheckAllFilled(self, storedValues, numberOfValues):
    def CheckAllFilled(self, storedValues, anyNan=True):        
        
        if storedValues.empty==False:
            if anyNan==True:
                ExistsNan=(pandas.isnull(storedValues.values)).any()
            if anyNan==False:
                #storedValues.iloc[[0],:]
                print(storedValues.iloc[:,[0]])
                print("storedValues.iloc[:,[0]]")
                print(pandas.isnull(storedValues.iloc[:,[0]]).values.any())
                print("pandas.isnull(storedValues.iloc[:,[0]]).values.any()")
                print(storedValues.iloc[:,[0]])
                
                print(len(storedValues.columns.values))
                print("len(storedValues.columns.values)")
                print("len(storedValues.column.values)")
                colnum=len(storedValues.columns.values)
                
                
                
                if ~(pandas.isnull(storedValues.iloc[:,[0]]).values.any()):
                    lisOfNan=[]
                    for i in range(1, colnum):
                        valueInList=~(pandas.isnull(storedValues.iloc[:,[i]]).values.all())
                        print(valueInList)
                        lisOfNan.append(valueInList)
                    print(lisOfNan)
                    if True in lisOfNan:
                        print("True in ListNan")
                        ExistsNan=False
                    else:
                        ExistsNan=True
                    
                else:
                   ExistsNan=True 
                    
                
# =============================================================================
#                 if pandas.isnull(storedValues.iloc[[0],:]).any().any():
#                     lisOfNan=[]
#                     for i in range(1,len(storedValues[0])):
#                         valueInList=pandas.isnull(storedValues.iloc[[i],:]).any().any()
#                         lisOfNan.append(valueInList)
#                     print(lisOfNan)    
#                     ExistsNan=False
# =============================================================================
        if storedValues.empty ==True:
            
            ExistsNan=True
# =============================================================================
#         if storedValues.empty ==False:
#             ExistsNan=False
#             for values in range(numberOfValues):
#                 step=str(storedValues.values[0][values])
#                 if step=='nan':
#                     ExistsNan=True
#         else:
#             ExistsNan=True
# =============================================================================
        print(ExistsNan)   
        return ExistsNan
    
    
    
    def NameFileToSaveAs(self):
        label= "What is the filename to save as? "
        WindowTitle= "Save As?"
        
        rsp, filename=self.JobNameWindow(JobName=False, label=label, WindowTitle=WindowTitle)
        
        return rsp, filename
    
    
    def NewJobNameWindow(self):
        label= "Name the new data fill job bellow:"
        WindowTitle= "Name The New Data Fill Job."
        
        rsp, filename=self.JobNameWindow(JobName=False, label=label, WindowTitle=WindowTitle)
        
        return rsp, filename
        
    
    


    def JobNameWindow(self, JobName=True, label=None, WindowTitle=None):
        DataFillName = QtWidgets.QDialog()
        self.ui = Ui_DataFillName()
        self.ui.setupUi(DataFillName)
        
        if JobName==False:
            self.ui.label.setText(label)
            DataFillName.setWindowTitle(WindowTitle)
        
        
        DataFillName.show()
        
        rsp=DataFillName.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp , self.ui.DataFIllName.text()
        else:
            return rsp , self.ui.DataFIllName.text()
        
        
        
    def clickedCopyMark(self):
        print("copy button clicked")
        cb = QtWidgets.QApplication.clipboard()
        cb.clear(mode=cb.Clipboard )
        cb.setText("XYXYXYXYX", mode=cb.Clipboard)
        
        




    def ChangeStartupDirectory(self, Folder):
        
        curentWorkingDirectory = os.getcwd()
        print(curentWorkingDirectory)
        print("curentWorkingDirectory")
        
        home = os.path.expanduser('~')
        
        print(home)
        
        location = os.path.join(home, 'Documents', Folder)
        
        print(location)
        
        folder_check = os.path.isdir(location)
        
        print(folder_check)
        
        if not os.path.exists(location):
            os.makedirs(location)
            print('folder created')
        else:
            print("folder exists")
         
        os.chdir(location)
        
        curentWorkingDirectory = os.getcwd()
        print(curentWorkingDirectory)
        print("New curentWorkingDirectory")
        
        
    def Inframe(self, frame):
        if '' in frame:
            ExistsNan=True
        else:
            ExistsNan=False
        return ExistsNan
    
    def writeToSqlite(self,frame,table_name):
        conn = sqlite3.connect(self.DatabaseName)
        print(frame)
        frame.to_sql(table_name, conn, if_exists="replace")

    def readSqlDatabase(self,table_name,columns=None):
        conn = create_engine('sqlite:///'+self.DatabaseName)
        try:
            df2=pandas.read_sql_table(table_name, conn, columns=columns)
            #print(df2)
            TableFound=True
        except ValueError:
            df2=None
            TableFound=False
        return df2,TableFound
        
        
        
    def OperaConcluidaWindow(self):
        
        OperaConcluidaDialog = QtWidgets.QDialog()
        ui = Ui_OperaConcluidaDialog()
        ui.setupUi(OperaConcluidaDialog)
        OperaConcluidaDialog.show()
        
        rsp=OperaConcluidaDialog.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        
    def CopyPasteWindow(self):
        
        CopyPasteMark = QtWidgets.QDialog()
        self.ui3 = Ui_CopyPasteMark()
        self.ui3.setupUi(CopyPasteMark)
        
        self.clickedCopyMark()
        
        self.ui3.CopyMarkButton.clicked.connect(self.clickedCopyMark)
        
        CopyPasteMark.show()
        
        rsp=CopyPasteMark.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass
    
    def InicioSalveDocumento(self):
        
        DConfeINICIALSalvo = QtWidgets.QDialog()
        ui = Ui_DConfeINICIALSalvo()
        ui.setupUi(DConfeINICIALSalvo)
        DConfeINICIALSalvo.show()
        
        rsp=DConfeINICIALSalvo.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass


    def addInformationANDSave(self):
        DAddInformation = QtWidgets.QDialog()
        ui = Ui_DAddInformation()
        ui.setupUi(DAddInformation)
        DAddInformation.show()
        
        rsp=DAddInformation.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass
    
    def fileAlreadyExits(self):
        DFileExists = QtWidgets.QDialog()
        ui = Ui_DFileExists()
        ui.setupUi(DFileExists)
        DFileExists.show()
        
        rsp=DFileExists.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass
    
    
    def addInfoTable(self, frame, columnNames):
        DInfoTable = QtWidgets.QDialog()
        self.ui4 = Ui_DInfoTable()
        self.ui4.setupUi(DInfoTable)
        DInfoTable.show()
        
        
        self.ui4.tableWidget.setColumnCount(0)
        numColumn=len(columnNames)
        self.ui4.tableWidget.setColumnCount(numColumn)
        
        
        print(columnNames)
        print("columnNames in for loop")
        
        self.ui4.tableWidget.setHorizontalHeaderLabels(columnNames)
        for i in range(len(columnNames)):
            self.ui4.tableWidget.setItem(0, i, QTableWidgetItem(frame[i]))
        
        
        rsp=DInfoTable.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            frame=[]
            for i in range(len(columnNames)):
                print(self.ui4.tableWidget.item(0,i).text())
                CellValue=self.ui4.tableWidget.item(0,i).text()
                
                CellValue=CellValue.split()
                CellValue=" ".join(CellValue)
                
                frame.append(CellValue)
                
            print(frame)
            
            return rsp, frame
        else:
            return rsp, frame
        
        pass
    
    def FileCanNotStartWith(self):
        rsp=self.AddInfoEXCEL(FileNameWithout=True)
        
        return rsp
        
    
    
    def AddInfoEXCEL(self, FileNameWithout=False):
        DAddInformation = QtWidgets.QDialog()
        self.ui2 = Ui_DAddInformation()
        self.ui2.setupUi(DAddInformation)
        
        if FileNameWithout:
            labelUsed="File chosen can not start with 'temp_'. Click OK then choose a different file to Continue or click CANCEL to cancel the operation."
            WindowTitle="File chosen can not start with 'temp_'!"
            self.ui2.label.setText(labelUsed)
            DAddInformation.setWindowTitle(WindowTitle)
        
        
        DAddInformation.show()
        
        
        rsp=DAddInformation.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass    
    
    
    def DJobNameNotExit(self):
        DJobNameNotExist = QtWidgets.QDialog()
        ui = Ui_DJobNameNotExist()
        ui.setupUi(DJobNameNotExist)
        DJobNameNotExist.show()
        
        rsp=DJobNameNotExist.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        pass 
    
    
    def DloadItens(self, datafillName):
        rsp=self.DAreYouSureDelete(datafillName=datafillName,WindowTitle="Load the item?", label="Do you want to load the item below?")
        return rsp
    
    def DAreYouSureDelete(self, datafillName,WindowTitle="DELETE THE ITEM?", label="DO YOU WANT TO DELETE THE ITEM BELOW?"):
        DDeleteItemCheck = QtWidgets.QDialog()
        self.ui5 = Ui_DDeleteItemCheck()
        self.ui5.setupUi(DDeleteItemCheck)
        
        DDeleteItemCheck.show()
        self.ui5.ItemText.setText(str(datafillName))
        
        DDeleteItemCheck.setWindowTitle(WindowTitle)
        
        self.ui5.label.setText(label)
                
        
        
        rsp=DDeleteItemCheck.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
    def DCloseDocToContinue(self, DocumentNameUsed):
        DcloseDoc2Continue = QtWidgets.QDialog()
        self.ui7 = Ui_DcloseDoc2Continue()
        self.ui7.setupUi(DcloseDoc2Continue)
        DcloseDoc2Continue.show()
        
        self.ui7.DocumentName.setText(str(DocumentNameUsed))
        
        
        rsp=DcloseDoc2Continue.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
    def DProriFileNamesUsed(self,filenameUsed, FolderUsed=False, SingleFileNameUsed=False, JobNameRedoOptions=False):
        DProhibFileNames = QtWidgets.QDialog()
        self.ui8 = Ui_DProhibFileNames()
        self.ui8.setupUi(DProhibFileNames)
        
        self.IgonoreClicked=False
        
        DProhibFileNames.show()
        
        
        
        self.ui8.filenames.setText(str(filenameUsed))
        if FolderUsed==True:
            DProhibFileNames.resize(470, 160)
            label="The folder chosen has the prohibited file name(s) listed below. Choose a different folder to continue. If you desire to use the folder and all its information, change the filename(s) in the folder click OK and re-choose this folder. Click CANCEL to cancel the operation. Click IGNORE to continue with this folder ignoring the listed filename(s) and the contained information."
            WindowTitle="The folder chosen has prohibited file name(s)!"
            self.ui8.label.setText(label)
            DProhibFileNames.setWindowTitle(WindowTitle)
        if SingleFileNameUsed==True:
            label="Filename already exists in the database. Click OK and change the filename used to a filename not registered, click CANCEL to cancel the process, or click Rewrite to rewrite the file."
            WindowTitle="Filename already exists in the database!"
            ButtonText="Rewrite"
            self.ui8.label.setText(label)
            DProhibFileNames.setWindowTitle(WindowTitle)
            self.ui8.BIgnore.setText(ButtonText)
            
        if JobNameRedoOptions:
            label="Data fill job name already exists in the database. Click OK and change the job name used to a job name not registered, click CANCEL to cancel the process, or click Rewrite to rewrite the job name."
            WindowTitle="Data fill job name already exists in the database!"
            ButtonText="Rewrite"
            self.ui8.label.setText(label)
            DProhibFileNames.setWindowTitle(WindowTitle)
            self.ui8.BIgnore.setText(ButtonText)
            
            
            pass
            
        
        self.ui8.BIgnore.clicked.connect(self.FuncIgnore)
        
        rsp=DProhibFileNames.exec_()
        
        if rsp==QtWidgets.QDialog.Accepted:
            return rsp 
        else:
            return rsp 
        
        
    def DProriFileNamesInFolderUsed(self,filenameUsed):
        rsp=self.DProriFileNamesUsed(filenameUsed=filenameUsed,FolderUsed=True)
        
        return rsp
    
    def DSingleFileNameUsed(self,filenameUsed):
        rsp=self.DProriFileNamesUsed(filenameUsed=filenameUsed,SingleFileNameUsed=True)
        
        return rsp
    
    def DRewriteDatafillJobName(self,filenameUsed):
        rsp=self.DProriFileNamesUsed(filenameUsed=filenameUsed,JobNameRedoOptions=True)
        
        return rsp




if __name__ == "__main__":

    if not QtWidgets.QApplication.instance():
        app = QtWidgets.QApplication(sys.argv)
    else:
        app = QtWidgets.QApplication.instance() 

    autoFormFiller()
    
    #autoFormFiller()
                
    					

